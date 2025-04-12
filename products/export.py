import csv
import io
import datetime
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def export_products_csv(queryset):
    """Export products to CSV format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products_export_{}.csv"'.format(
        datetime.datetime.now().strftime('%Y-%m-%d')
    )
    
    writer = csv.writer(response)
    # Write header row
    writer.writerow(['Name', 'SKU', 'Category', 'Price', 'Cost', 'Quantity', 
                    'Minimum Stock', 'Status', 'Supplier', 'Created At'])
    
    # Write data rows
    for product in queryset:
        writer.writerow([
            product.name,
            product.sku,
            product.category.name if product.category else '',
            product.price,
            product.cost,
            product.quantity,
            product.minimum_stock,
            product.status,
            product.supplier,
            product.created_at.strftime('%Y-%m-%d %H:%M')
        ])
    
    return response

def export_products_excel(queryset):
    """Export products to Excel format"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="products_export_{}.xlsx"'.format(
        datetime.datetime.now().strftime('%Y-%m-%d')
    )
    
    # Create a workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Products"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write header row
    headers = ['Name', 'SKU', 'Category', 'Price', 'Cost', 'Quantity', 
              'Minimum Stock', 'Status', 'Supplier', 'Created At']
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data rows
    for row_num, product in enumerate(queryset, 2):
        worksheet.cell(row=row_num, column=1).value = product.name
        worksheet.cell(row=row_num, column=2).value = product.sku
        worksheet.cell(row=row_num, column=3).value = product.category.name if product.category else ''
        worksheet.cell(row=row_num, column=4).value = float(product.price)
        worksheet.cell(row=row_num, column=5).value = float(product.cost) if product.cost else 0
        worksheet.cell(row=row_num, column=6).value = product.quantity
        worksheet.cell(row=row_num, column=7).value = product.minimum_stock
        worksheet.cell(row=row_num, column=8).value = product.status
        worksheet.cell(row=row_num, column=9).value = product.supplier
        worksheet.cell(row=row_num, column=10).value = product.created_at.strftime('%Y-%m-%d %H:%M')
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())
    
    return response

def export_products_pdf(queryset):
    """Export products to PDF format"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="products_export_{}.pdf"'.format(
        datetime.datetime.now().strftime('%Y-%m-%d')
    )
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    
    # Add title
    elements.append(Paragraph("Products Export", title_style))
    elements.append(Spacer(1, 20))
    
    # Prepare data for table
    data = [['Name', 'SKU', 'Category', 'Price', 'Cost', 'Quantity', 
             'Min Stock', 'Status', 'Supplier', 'Created At']]
    
    for product in queryset:
        data.append([
            product.name,
            product.sku,
            product.category.name if product.category else '',
            str(product.price),
            str(product.cost) if product.cost else '0',
            str(product.quantity),
            str(product.minimum_stock),
            product.status,
            product.supplier,
            product.created_at.strftime('%Y-%m-%d')
        ])
    
    # Create table
    table = Table(data)
    
    # Style the table
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    
    # Add alternating row colors
    for i in range(1, len(data)):
        if i % 2 == 0:
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.lightgrey)
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response

def export_categories_csv(queryset):
    """Export categories to CSV format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="categories_export_{}.csv"'.format(
        datetime.datetime.now().strftime('%Y-%m-%d')
    )
    
    writer = csv.writer(response)
    # Write header row
    writer.writerow(['Name', 'Description', 'Product Count', 'Created At'])
    
    # Write data rows
    for category in queryset:
        writer.writerow([
            category.name,
            category.description or '',
            category.products.count(),
            category.created_at.strftime('%Y-%m-%d %H:%M')
        ])
    
    return response

def export_categories_excel(queryset):
    """Export categories to Excel format"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="categories_export_{}.xlsx"'.format(
        datetime.datetime.now().strftime('%Y-%m-%d')
    )
    
    # Create a workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Categories"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write header row
    headers = ['Name', 'Description', 'Product Count', 'Created At']
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data rows
    for row_num, category in enumerate(queryset, 2):
        worksheet.cell(row=row_num, column=1).value = category.name
        worksheet.cell(row=row_num, column=2).value = category.description or ''
        worksheet.cell(row=row_num, column=3).value = category.products.count()
        worksheet.cell(row=row_num, column=4).value = category.created_at.strftime('%Y-%m-%d %H:%M')
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())
    
    return response

def export_categories_pdf(queryset):
    """Export categories to PDF format"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="categories_export_{}.pdf"'.format(
        datetime.datetime.now().strftime('%Y-%m-%d')
    )
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    
    # Add title
    elements.append(Paragraph("Categories Export", title_style))
    elements.append(Spacer(1, 20))
    
    # Prepare data for table
    data = [['Name', 'Description', 'Product Count', 'Created At']]
    
    for category in queryset:
        data.append([
            category.name,
            category.description or '',
            str(category.products.count()),
            category.created_at.strftime('%Y-%m-%d')
        ])
    
    # Create table
    table = Table(data)
    
    # Style the table
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    
    # Add alternating row colors
    for i in range(1, len(data)):
        if i % 2 == 0:
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.lightgrey)
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response
